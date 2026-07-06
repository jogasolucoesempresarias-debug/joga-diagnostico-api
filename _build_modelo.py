"""
Gera templates/modelo_diagnostico.xlsx — planilha-modelo (fallback camada 3).
Rodar uma vez: python _build_modelo.py
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEST = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates", "modelo_diagnostico.xlsx")

_HEADER_FILL = PatternFill("solid", fgColor="13171E")
_HEADER_FONT = Font(color="F4A52A", bold=True)


def _aba(ws, colunas, exemplos):
    ws.append(colunas)
    for cel in ws[1]:
        cel.fill = _HEADER_FILL
        cel.font = _HEADER_FONT
    for linha in exemplos:
        ws.append(linha)
    for i, _ in enumerate(colunas, start=1):
        ws.column_dimensions[chr(64 + i)].width = 22


def main():
    wb = Workbook()

    carteira = wb.active
    carteira.title = "Carteira"
    _aba(
        carteira,
        ["cliente", "data", "valor"],
        [
            ["Cliente Exemplo A", "01/03/2026", "1500,00"],
            ["Cliente Exemplo A", "20/04/2026", "980,50"],
            ["Cliente Exemplo B", "15/02/2026", "3200,00"],
        ],
    )

    estoque = wb.create_sheet("Estoque")
    _aba(
        estoque,
        ["produto", "saldo", "custo", "venda_periodo"],
        [
            ["Produto Exemplo 1", "120", "18,50", "40"],
            ["Produto Exemplo 2", "5", "9,90", "60"],
        ],
    )

    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    wb.save(DEST)
    print(f"OK: {DEST}")


if __name__ == "__main__":
    main()
